#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Yeniden Değerleme ve Amortisman - Web Arayüzü
"""

from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import tempfile
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

# Mevcut fonksiyonlar
def hesapla_yd_orani(aktif_tarihi, islem_yili, defter_net_degeri, genel_yd_orani):
    if defter_net_degeri == 0:
        return 0
    aktif_yili = aktif_tarihi.year
    if aktif_yili == islem_yili:
        return 0
    return genel_yd_orani

def hesapla_yillik_amortisman(yd_sabit_kiymet, yd_net_deger, amortisman_orani, yontem, defter_net_degeri, aktif_tarihi, islem_yili):
    if defter_net_degeri == 0 or yd_net_deger == 0:
        return 0
    yontem = yontem.strip() if isinstance(yontem, str) else str(yontem)
    if yontem == "Normal":
        return yd_sabit_kiymet * amortisman_orani
    elif yontem == "Hızlı":
        # Hızlı amortisman yöntemi için son yıl kontrolü
        aktif_yili = aktif_tarihi.year
        gecen_yil_sayisi = islem_yili - aktif_yili

        # Amortisman oranından faydalı ömrü hesapla (örn: %20 = 5 yıl)
        faydali_omur = int(1 / amortisman_orani) if amortisman_orani > 0 else 0

        # Son yıl mı kontrol et (son yıl = faydalı ömür - 1)
        if faydali_omur > 0 and gecen_yil_sayisi >= (faydali_omur - 1):
            # Son yılda kalan bakiyenin tamamını amortisman olarak ayır
            return yd_net_deger
        else:
            # Normal hızlı amortisman hesaplaması
            return yd_net_deger * (amortisman_orani * 2)
    else:
        return yd_sabit_kiymet * amortisman_orani

def hesapla_donem_amortismani(yillik_amortisman, donem):
    if donem == 1:
        return yillik_amortisman / 4
    elif donem == 2:
        return yillik_amortisman / 2
    elif donem == 3:
        return (yillik_amortisman / 4) * 3
    else:
        return yillik_amortisman

def sabit_kiymet_okuma_from_file(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    sabit_kiymetler = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break

        # Değerleri sayıya çevir
        try:
            defter_son_degeri = float(row[5]) if row[5] is not None else 0
        except (ValueError, TypeError):
            defter_son_degeri = 0

        try:
            defter_birikmis_amort = float(row[6]) if row[6] is not None else 0
        except (ValueError, TypeError):
            defter_birikmis_amort = 0

        # Net değeri hesapla (8. kolon formül veya sayı olabilir)
        if row[7] is not None:
            try:
                defter_net_degeri = float(row[7])
            except (ValueError, TypeError):
                defter_net_degeri = defter_son_degeri - defter_birikmis_amort
        else:
            defter_net_degeri = defter_son_degeri - defter_birikmis_amort

        # Amortisman oranını düzelt (yüzde formatında ise ondalığa çevir)
        try:
            amortisman_orani = float(row[3]) if row[3] else 0
            # Eğer 1'den büyükse yüzde formatında yazılmış demektir (örn: 20 = %20)
            if amortisman_orani > 1:
                amortisman_orani = amortisman_orani / 100
        except (ValueError, TypeError):
            amortisman_orani = 0

        sabit_kiymet = {
            'hesap_kodu': row[0],
            'aciklama': row[1],
            'aktif_tarihi': row[2],
            'amortisman_orani': amortisman_orani,
            'amortisman_yontemi': row[4],
            'defter_son_degeri': defter_son_degeri,
            'defter_birikmis_amort': defter_birikmis_amort,
            'defter_net_degeri': defter_net_degeri
        }
        sabit_kiymetler.append(sabit_kiymet)

    wb.close()
    return sabit_kiymetler

def yeniden_degerleme_hesapla(sabit_kiymetler, islem_yili, donem, yd_orani_genel):
    sonuclar = []
    for sk in sabit_kiymetler:
        yd_orani = hesapla_yd_orani(
            sk['aktif_tarihi'], islem_yili, sk['defter_net_degeri'], yd_orani_genel
        )
        yd_sabit_kiymet = sk['defter_son_degeri'] * (1 + yd_orani)
        yd_birikmis_amort = sk['defter_birikmis_amort'] * (1 + yd_orani)
        yd_net_deger = yd_sabit_kiymet - yd_birikmis_amort
        yd_yillik_amortisman = hesapla_yillik_amortisman(
            yd_sabit_kiymet, yd_net_deger, sk['amortisman_orani'],
            sk['amortisman_yontemi'], sk['defter_net_degeri'],
            sk['aktif_tarihi'], islem_yili
        )
        yd_donem_amortismani = hesapla_donem_amortismani(yd_yillik_amortisman, donem)

        sonuc = {
            **sk,
            'yd_orani': yd_orani,
            'yd_sabit_kiymet': yd_sabit_kiymet,
            'yd_birikmis_amort': yd_birikmis_amort,
            'yd_net_deger': yd_net_deger,
            'yd_yillik_amortisman': yd_yillik_amortisman,
            'yd_donem_amortismani': yd_donem_amortismani
        }
        sonuclar.append(sonuc)

    return sonuclar

def muhasebe_fisleri_olustur(sonuclar):
    fisler = {'yeniden_degerleme': [], 'amortisman': []}
    hesap_gruplari = {}

    for sonuc in sonuclar:
        hesap_kodu = sonuc['hesap_kodu']
        if hesap_kodu not in hesap_gruplari:
            hesap_gruplari[hesap_kodu] = []
        hesap_gruplari[hesap_kodu].append(sonuc)

    for hesap_kodu, grup in hesap_gruplari.items():
        toplam_defter_son = sum(s['defter_son_degeri'] for s in grup)
        toplam_yd_sabit = sum(s['yd_sabit_kiymet'] for s in grup)
        toplam_defter_birikmis = sum(s['defter_birikmis_amort'] for s in grup)
        toplam_yd_birikmis = sum(s['yd_birikmis_amort'] for s in grup)

        fark_sabit = toplam_yd_sabit - toplam_defter_son
        fark_birikmis = toplam_yd_birikmis - toplam_defter_birikmis
        fark_net = fark_sabit - fark_birikmis

        if abs(fark_sabit) > 0.01 or abs(fark_birikmis) > 0.01:
            birikmis_amort_hesap = 257 if str(hesap_kodu).startswith('25') else 268
            fis = {
                'aciklama': f'Yeniden Değerleme - {hesap_kodu}',
                'kayitlar': [
                    {'hesap': hesap_kodu, 'borc': round(fark_sabit, 2), 'alacak': 0},
                    {'hesap': birikmis_amort_hesap, 'borc': 0, 'alacak': round(fark_birikmis, 2)},
                    {'hesap': 522, 'borc': 0, 'alacak': round(fark_net, 2)}
                ]
            }
            fisler['yeniden_degerleme'].append(fis)

    for hesap_kodu, grup in hesap_gruplari.items():
        toplam_donem_amort = sum(s['yd_donem_amortismani'] for s in grup)
        if abs(toplam_donem_amort) > 0.01:
            birikmis_amort_hesap = 257 if str(hesap_kodu).startswith('25') else 268
            fis = {
                'aciklama': f'Dönem Amortismanı - {hesap_kodu}',
                'kayitlar': [
                    {'hesap': 770, 'borc': round(toplam_donem_amort, 2), 'alacak': 0},
                    {'hesap': birikmis_amort_hesap, 'borc': 0, 'alacak': round(toplam_donem_amort, 2)}
                ]
            }
            fisler['amortisman'].append(fis)

    return fisler

def excel_dosyasi_olustur(sonuclar, fisler, islem_yili, donem_adi, yd_orani_genel):
    wb = openpyxl.Workbook()
    baslik_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    baslik_font = Font(bold=True, color="FFFFFF", size=11)
    yd_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    yd_font = Font(bold=True, size=10)
    toplam_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    toplam_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws1 = wb.active
    ws1.title = "YD ve Amortisman"

    ws1['A1'] = 'YENİDEN DEĞERLEME VE AMORTİSMAN TABLOSU'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:P1')
    ws1['A2'] = f'İşlem Yılı: {islem_yili}'
    ws1['A2'].font = Font(bold=True, size=11)
    ws1['A3'] = f'Dönem: {donem_adi}'
    ws1['A3'].font = Font(bold=True, size=11)
    ws1['A4'] = f'YD Oranı: %{yd_orani_genel * 100:.4f}'
    ws1['A4'].font = Font(bold=True, size=11)

    basliklar = [
        'Sabit Kıymet', 'Açıklama', 'Aktif Giriş Tarihi', 'Amort. Oranı',
        'Amort. Yöntemi', 'Defter Son Değeri', 'Defter Birikmiş Amort.',
        'Defter Net Değeri', '', 'YD Oranı', 'YD Sabit Kıymet',
        'YD Birikmiş Amort.', 'YD Net Değer', '', 'YD Yıllık Amortisman',
        'YD Dönem Amortismanı'
    ]

    for col_num, baslik in enumerate(basliklar, 1):
        cell = ws1.cell(row=6, column=col_num)
        cell.value = baslik
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row_num = 7
    for sonuc in sonuclar:
        ws1.cell(row=row_num, column=1).value = sonuc['hesap_kodu']
        ws1.cell(row=row_num, column=2).value = sonuc['aciklama']
        ws1.cell(row=row_num, column=3).value = sonuc['aktif_tarihi'].strftime('%d.%m.%Y')
        ws1.cell(row=row_num, column=4).value = sonuc['amortisman_orani']
        ws1.cell(row=row_num, column=4).number_format = '0.00%'
        ws1.cell(row=row_num, column=5).value = sonuc['amortisman_yontemi']
        ws1.cell(row=row_num, column=6).value = sonuc['defter_son_degeri']
        ws1.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=7).value = sonuc['defter_birikmis_amort']
        ws1.cell(row=row_num, column=7).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=8).value = sonuc['defter_net_degeri']
        ws1.cell(row=row_num, column=8).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=10).value = sonuc['yd_orani']
        ws1.cell(row=row_num, column=10).number_format = '0.00%'
        ws1.cell(row=row_num, column=10).fill = yd_fill
        ws1.cell(row=row_num, column=11).value = sonuc['yd_sabit_kiymet']
        ws1.cell(row=row_num, column=11).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=11).fill = yd_fill
        ws1.cell(row=row_num, column=12).value = sonuc['yd_birikmis_amort']
        ws1.cell(row=row_num, column=12).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=12).fill = yd_fill
        ws1.cell(row=row_num, column=13).value = sonuc['yd_net_deger']
        ws1.cell(row=row_num, column=13).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=13).fill = yd_fill
        ws1.cell(row=row_num, column=15).value = sonuc['yd_yillik_amortisman']
        ws1.cell(row=row_num, column=15).number_format = '#,##0.00'
        ws1.cell(row=row_num, column=16).value = sonuc['yd_donem_amortismani']
        ws1.cell(row=row_num, column=16).number_format = '#,##0.00'

        for col in range(1, 17):
            ws1.cell(row=row_num, column=col).border = border
        row_num += 1

    ws1.cell(row=row_num, column=2).value = 'TOPLAM'
    ws1.cell(row=row_num, column=2).font = toplam_font
    ws1.cell(row=row_num, column=2).fill = toplam_fill

    baslangic_row = 7
    bitis_row = row_num - 1
    toplam_kolonlar = [6, 7, 8, 11, 12, 13, 15, 16]
    for col in toplam_kolonlar:
        cell = ws1.cell(row=row_num, column=col)
        cell.value = f'=SUM({cell.column_letter}{baslangic_row}:{cell.column_letter}{bitis_row})'
        cell.number_format = '#,##0.00'
        cell.font = toplam_font
        cell.fill = toplam_fill
        cell.border = border

    for col in range(1, 17):
        ws1.cell(row=row_num, column=col).border = border

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 18
    ws1.column_dimensions['H'].width = 18
    ws1.column_dimensions['I'].width = 2
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 18
    ws1.column_dimensions['L'].width = 18
    ws1.column_dimensions['M'].width = 18
    ws1.column_dimensions['N'].width = 2
    ws1.column_dimensions['O'].width = 18
    ws1.column_dimensions['P'].width = 18

    ws2 = wb.create_sheet(title="Muhasebe Fişleri")
    ws2['A1'] = 'MUHASEBE FİŞLERİ'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:D1')

    row_num = 3
    ws2.cell(row=row_num, column=1).value = 'YENİDEN DEĞERLEME FİŞLERİ'
    ws2.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_num, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2.merge_cells(f'A{row_num}:D{row_num}')
    row_num += 1

    for fis in fisler['yeniden_degerleme']:
        ws2.cell(row=row_num, column=1).value = fis['aciklama']
        ws2.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        ws2.cell(row=row_num, column=1).value = 'Hesap Kodu'
        ws2.cell(row=row_num, column=2).value = 'Hesap Adı'
        ws2.cell(row=row_num, column=3).value = 'Borç'
        ws2.cell(row=row_num, column=4).value = 'Alacak'
        for col in range(1, 5):
            ws2.cell(row=row_num, column=col).font = Font(bold=True)
            ws2.cell(row=row_num, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws2.cell(row=row_num, column=col).border = border
        row_num += 1

        for kayit in fis['kayitlar']:
            ws2.cell(row=row_num, column=1).value = kayit['hesap']
            ws2.cell(row=row_num, column=2).value = ''
            ws2.cell(row=row_num, column=3).value = kayit['borc'] if kayit['borc'] > 0 else ''
            ws2.cell(row=row_num, column=3).number_format = '#,##0.00'
            ws2.cell(row=row_num, column=4).value = kayit['alacak'] if kayit['alacak'] > 0 else ''
            ws2.cell(row=row_num, column=4).number_format = '#,##0.00'
            for col in range(1, 5):
                ws2.cell(row=row_num, column=col).border = border
            row_num += 1
        row_num += 1

    row_num += 1
    ws2.cell(row=row_num, column=1).value = 'AMORTİSMAN FİŞLERİ'
    ws2.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws2.cell(row=row_num, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2.merge_cells(f'A{row_num}:D{row_num}')
    row_num += 1

    for fis in fisler['amortisman']:
        ws2.cell(row=row_num, column=1).value = fis['aciklama']
        ws2.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        ws2.cell(row=row_num, column=1).value = 'Hesap Kodu'
        ws2.cell(row=row_num, column=2).value = 'Hesap Adı'
        ws2.cell(row=row_num, column=3).value = 'Borç'
        ws2.cell(row=row_num, column=4).value = 'Alacak'
        for col in range(1, 5):
            ws2.cell(row=row_num, column=col).font = Font(bold=True)
            ws2.cell(row=row_num, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws2.cell(row=row_num, column=col).border = border
        row_num += 1

        for kayit in fis['kayitlar']:
            ws2.cell(row=row_num, column=1).value = kayit['hesap']
            ws2.cell(row=row_num, column=2).value = ''
            ws2.cell(row=row_num, column=3).value = kayit['borc'] if kayit['borc'] > 0 else ''
            ws2.cell(row=row_num, column=3).number_format = '#,##0.00'
            ws2.cell(row=row_num, column=4).value = kayit['alacak'] if kayit['alacak'] > 0 else ''
            ws2.cell(row=row_num, column=4).number_format = '#,##0.00'
            for col in range(1, 5):
                ws2.cell(row=row_num, column=col).border = border
            row_num += 1
        row_num += 1

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    tarih_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    dosya_adi = f'YD_Amortisman_Sonuc_{islem_yili}_{donem_adi.replace(" ", "_").replace(".", "")}_{tarih_str}.xlsx'
    dosya_yolu = os.path.join(app.config['UPLOAD_FOLDER'], dosya_adi)
    wb.save(dosya_yolu)

    return dosya_yolu, dosya_adi

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/hesapla', methods=['POST'])
def hesapla():
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'Excel dosyası yüklenmedi'}), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'Dosya seçilmedi'}), 400

        islem_yili = int(request.form['islem_yili'])
        donem = int(request.form['donem'])
        yd_orani = float(request.form['yd_orani'].replace(',', '.')) / 100

        donem_adi = {1: "1. Dönem", 2: "2. Dönem", 3: "3. Dönem", 4: "Yıllık"}[donem]

        filename = secure_filename(file.filename)
        temp_input = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(temp_input)

        sabit_kiymetler = sabit_kiymet_okuma_from_file(temp_input)
        sonuclar = yeniden_degerleme_hesapla(sabit_kiymetler, islem_yili, donem, yd_orani)
        fisler = muhasebe_fisleri_olustur(sonuclar)
        dosya_yolu, dosya_adi = excel_dosyasi_olustur(sonuclar, fisler, islem_yili, donem_adi, yd_orani)

        os.remove(temp_input)

        return jsonify({
            'success': True,
            'message': 'Hesaplama başarıyla tamamlandı!',
            'download_url': f'/indir/{dosya_adi}',
            'sabit_kiymet_sayisi': len(sonuclar),
            'yd_fis_sayisi': len(fisler['yeniden_degerleme']),
            'amortisman_fis_sayisi': len(fisler['amortisman'])
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/indir/<filename>')
def indir(filename):
    dosya_yolu = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(dosya_yolu, as_attachment=True, download_name=filename)

@app.route('/sablon-indir')
def sablon_indir():
    """Şablon Excel dosyasını oluştur ve indir"""
    from datetime import datetime

    # Yeni workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Şablon"

    # Stil tanımlamaları
    baslik_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    baslik_font = Font(bold=True, color="FFFFFF", size=12)
    ornek_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Başlık satırı
    basliklar = [
        'sabit kıymet', 'sabit kıymet açıklama', 'aktife giriş tarihi',
        'amortisman oranı', 'amortisman yöntemi', 'defter son değeri',
        'defter birikmiş amort', 'defter net değeri'
    ]

    for col_num, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = baslik
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Örnek veri satırları
    ornekler = [
        [254, 'BMW Araç', datetime(2024, 1, 1), 0.20, 'Normal', 3000000, 600000, '=F2-G2'],
        [254, 'Mercedes Araç', datetime(2025, 1, 1), 0.20, 'Normal', 2000000, 0, '=F3-G3'],
        [255, 'Daktilo', datetime(2023, 12, 1), 0.10, 'Hızlı', 1000000, 190000, '=F4-G4'],
        [253, 'Makine', datetime(2020, 10, 1), 0.20, 'Normal', 5000000, 5000000, '=F5-G5'],
    ]

    for row_num, ornek in enumerate(ornekler, 2):
        for col_num, value in enumerate(ornek, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if col_num == 3:  # Tarih
                cell.value = value
                cell.number_format = 'DD.MM.YYYY'
            elif col_num == 4:  # Amortisman oranı
                cell.value = value
                cell.number_format = '0.00'
            elif col_num in [6, 7]:  # Tutarlar
                cell.value = value
                cell.number_format = '#,##0.00'
            else:
                cell.value = value
            cell.fill = ornek_fill
            cell.border = border

    # Kolon genişlikleri
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18

    # Geçici dosyaya kaydet
    temp_path = os.path.join(tempfile.gettempdir(), 'SABLON_SABIT_KIYMET_LISTESI.xlsx')
    wb.save(temp_path)

    return send_file(temp_path, as_attachment=True, download_name='SABLON_SABIT_KIYMET_LISTESI.xlsx')

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 8080))
    debug = os.environ.get('DEBUG', 'False') == 'True'

    if not debug:
        print("\n" + "="*60)
        print("YENİDEN DEĞERLEME VE AMORTİSMAN - WEB ARAYÜZÜ")
        print("="*60)
        print("\nTarayıcınızda şu adresi açın:")
        print(f"http://localhost:{port}")
        print("\nProgramı durdurmak için: CTRL+C")
        print("="*60 + "\n")

    app.run(debug=debug, host='0.0.0.0', port=port)
