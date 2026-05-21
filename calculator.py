from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MAX_UPLOAD_ROWS = 5000

COLUMN_ALIASES = {
    "kiymet_no": ["sabit kıymet", "sabit kiymet", "kiymet no", "demirbaş no", "demirbas no", "sıra no", "no", "kod"],
    "kiymet_ad": ["sabit kıymet açıklama", "sabit kiymet aciklama", "kıymet adı", "kiymet adi", "cinsi", "açıklama", "aciklama", "tanım", "tanim", "ad", "isim"],
    "aktif_hesap": ["aktif hesap", "hesap kodu", "hesap", "muhasebe kodu"],
    "gider_hesap": ["gider hesap", "gider hesap kodu", "gider", "gider kodu"],
    "tarih": ["aktife giriş tarihi", "aktife giris tarihi", "giriş tarihi", "giris tarihi", "edinim tarihi", "tarih", "alım tarihi", "alim tarihi", "fatura tarihi"],
    "maliyet": ["defter son değeri", "defter son degeri", "fiili maliyet", "maliyet", "maliyet bedeli", "bedel", "tutar", "değer", "deger", "fiyat"],
    "birikmis_amortisman": ["defter birikmiş amort", "defter birikmis amort", "birikmiş amortisman", "birikmis amortisman"],
    "net_deger": ["defter net değeri", "defter net degeri", "net değer", "net deger"],
    "amortisman_orani": ["amortisman oranı", "amortisman orani", "oran"],
    "omur": ["faydalı ömür", "faydali omur", "faydalı ömür yıl", "omur", "ömür", "süre", "sure", "yıl", "yil", "amortisman süresi"],
    "yontem": ["yöntem", "yontem", "amortisman yöntemi", "amort yontemi", "metot", "metod"],
    "binek": ["binek oto", "binek", "binek mi"],
}


@dataclass
class Asset:
    kiymet_no: str
    kiymet_ad: str
    aktif_hesap: str
    gider_hesap: str
    tarih: datetime
    maliyet: float
    omur: int
    yontem: str
    binek: str
    amortisman_orani: float
    birikmis_amortisman: float
    net_deger: float


def normalize(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).lower().strip()
    replacements = str.maketrans("çğıöşüİ", "cgiosui")
    value = value.translate(replacements)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def find_column(headers: list[Any], aliases: list[str]) -> int | None:
    normalized_headers = {normalize(header): idx for idx, header in enumerate(headers)}
    for alias in aliases:
        normalized_alias = normalize(alias)
        if normalized_alias in normalized_headers:
            return normalized_headers[normalized_alias]
        for normalized_header, idx in normalized_headers.items():
            if normalized_alias and (normalized_header.startswith(normalized_alias) or normalized_alias in normalized_header):
                return idx
    return None


def parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value))
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    return int(number) if number else None


def parse_yontem(value: Any) -> str:
    if value in (None, ""):
        return "Normal"
    text = normalize(value)
    return "Azalan" if "azal" in text or "hizli" in text or "declining" in text else "Normal"


def parse_binek(value: Any) -> str:
    if value in (None, ""):
        return "H"
    text = normalize(value).upper()
    return "E" if text in {"E", "EVET", "YES", "Y", "1", "TRUE", "X"} else "H"


def read_assets(path: str | Path) -> list[Asset]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel dosyasında veri bulunamadı.")

    header_row_index = _detect_header_row(rows)
    headers = list(rows[header_row_index])
    mapping = {field: find_column(headers, aliases) for field, aliases in COLUMN_ALIASES.items()}

    missing = [field for field in ("kiymet_ad", "tarih", "maliyet") if mapping[field] is None]
    if mapping["omur"] is None and mapping["amortisman_orani"] is None:
        missing.append("omur/amortisman_orani")
    if missing:
        raise ValueError("Zorunlu kolonlar bulunamadı: " + ", ".join(missing))

    assets: list[Asset] = []
    for index, row in enumerate(rows[header_row_index + 1 : header_row_index + 1 + MAX_UPLOAD_ROWS], start=1):
        asset = _asset_from_row(row, mapping, index)
        if asset:
            assets.append(asset)

    if not assets:
        raise ValueError("Hesaplanacak geçerli sabit kıymet satırı bulunamadı.")
    return assets


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:10]):
        headers = list(row)
        score = sum(1 for aliases in COLUMN_ALIASES.values() if find_column(headers, aliases) is not None)
        if score > best_score:
            best_index = idx
            best_score = score
    return best_index


def _asset_from_row(row: tuple[Any, ...], mapping: dict[str, int | None], index: int) -> Asset | None:
    def get(field: str, default: Any = None) -> Any:
        idx = mapping.get(field)
        return row[idx] if idx is not None and idx < len(row) else default

    name = get("kiymet_ad")
    date = parse_date(get("tarih"))
    cost = parse_number(get("maliyet"))
    rate = parse_number(get("amortisman_orani"))
    life = parse_int(get("omur"))
    if life is None and rate:
        life = max(round(1 / rate), 1)
    if not name or date is None or cost is None or life is None:
        return None
    accumulated = parse_number(get("birikmis_amortisman")) or 0
    net_value = parse_number(get("net_deger"))
    if net_value is None:
        net_value = cost - accumulated
    depreciation_rate = rate if rate else 1 / max(life, 1)

    return Asset(
        kiymet_no=str(get("kiymet_no", f"SK-{index:04d}") or f"SK-{index:04d}").strip(),
        kiymet_ad=str(name).strip(),
        aktif_hesap=str(get("aktif_hesap") or get("kiymet_no") or "255").strip(),
        gider_hesap=str(get("gider_hesap", "770") or "770").strip(),
        tarih=date,
        maliyet=cost,
        omur=max(life, 1),
        yontem=parse_yontem(get("yontem")),
        binek=parse_binek(get("binek")),
        amortisman_orani=depreciation_rate,
        birikmis_amortisman=accumulated,
        net_deger=net_value,
    )


def calculate_assets(assets: list[Asset], islem_yili: int, donem: int, yd_orani: float) -> list[dict[str, Any]]:
    period_months = {1: 3, 2: 6, 3: 9, 4: 12}.get(donem, 12)
    factor = yd_orani / 100
    results = []
    for asset in assets:
        eligible_for_revaluation = asset.tarih.year < islem_yili and asset.net_deger > 0
        asset_factor = factor if eligible_for_revaluation else 0
        is_passenger_car = _is_passenger_car(asset)
        active_months = _active_months(asset.tarih, islem_yili, period_months, asset.omur, is_passenger_car)
        status = "Amortisman hakkı yok" if active_months == 0 else _asset_status(asset, islem_yili, is_passenger_car)
        revalued_cost = asset.maliyet * (1 + asset_factor)
        revalued_accumulated = asset.birikmis_amortisman * (1 + asset_factor)
        revalued_net = revalued_cost - revalued_accumulated
        revaluation_increase = asset.maliyet * asset_factor
        accumulated_increase = asset.birikmis_amortisman * asset_factor
        fund_increase = asset.net_deger * asset_factor
        revalued_annual_depreciation = 0 if active_months == 0 else _annual_depreciation(asset, base=revalued_cost, net_base=revalued_net)
        revalued_period_depreciation = revalued_annual_depreciation * active_months / 12

        results.append(
            {
                "asset": asset,
                "eligible_for_revaluation": eligible_for_revaluation,
                "is_passenger_car": is_passenger_car,
                "active_months": active_months,
                "status": status,
                "yd_orani": yd_orani if eligible_for_revaluation else 0,
                "revalued_cost": revalued_cost,
                "revalued_accumulated": revalued_accumulated,
                "revalued_net": revalued_net,
                "revaluation_increase": revaluation_increase,
                "accumulated_increase": accumulated_increase,
                "fund_increase": fund_increase,
                "annual_depreciation": revalued_annual_depreciation,
                "period_depreciation": revalued_period_depreciation,
                "revalued_period_depreciation": revalued_period_depreciation,
            }
        )
    return results


def _active_months(date: datetime, year: int, period_months: int, useful_life: int, is_passenger_car: bool) -> int:
    if date.year > year:
        return 0
    if date.year < year:
        return 0 if year > date.year + useful_life - 1 else period_months
    if date.month > period_months:
        return 0
    if is_passenger_car:
        return period_months - date.month + 1
    return period_months


def _is_passenger_car(asset: Asset) -> bool:
    if asset.binek == "E":
        return True
    account = str(asset.aktif_hesap).strip()
    description = normalize(asset.kiymet_ad)
    return account == "254" or any(word in description for word in ("arac", "oto", "otomobil", "binek", "bmw", "mercedes", "audi"))


def _asset_status(asset: Asset, year: int, is_passenger_car: bool) -> str:
    last_year = asset.tarih.year + asset.omur - 1
    if year > last_year:
        return "Amortisman hakkı yok"
    if is_passenger_car:
        if year == last_year:
            return "Son yıl dikkat"
        if year == asset.tarih.year:
            return "Binek ilk yıl kıst"
        return ""
    return "Son yıl" if year == last_year else ""


def _annual_depreciation(asset: Asset, base: float | None = None, net_base: float | None = None) -> float:
    cost = asset.maliyet if base is None else base
    if net_base is not None and net_base <= 0:
        return 0
    normal = cost * asset.amortisman_orani
    if asset.yontem == "Azalan":
        declining_base = asset.net_deger if net_base is None else net_base
        return min(declining_base * asset.amortisman_orani * 2, cost * 0.5)
    return normal


def create_template(path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Şablon"
    headers = ["sabit kıymet", "sabit kıymet açıklama", "aktife giriş tarihi", "amortisman oranı", "amortisman yöntemi", "defter son değeri", "defter birikmiş amort", "defter net değeri"]
    examples = [
        [252, "Fabrika Binası", "01.01.2020", 0.02, "Normal", 12000000, 1440000, "=F2-G2"],
        [253, "CNC Makinesi", "15.03.2021", 0.20, "Hızlı", 3500000, 2450000, "=F3-G3"],
        [253, "Paketleme Hattı", "10.07.2023", 0.10, "Normal", 2800000, 560000, "=F4-G4"],
        [254, "BMW Araç", "01.11.2022", 0.20, "Normal", 2000000, 1400000, "=F5-G5"],
        [254, "Mercedes Araç", "01.04.2026", 0.20, "Normal", 1200000, 0, "=F6-G6"],
        [255, "Ofis Mobilyaları", "01.01.2022", 0.20, "Hızlı", 500000, 400000, "=F7-G7"],
        [255, "Bilgisayar Donanımı", "05.05.2024", 0.25, "Hızlı", 750000, 375000, "=F8-G8"],
        [264, "Özel Maliyet", "01.09.2023", 0.20, "Normal", 1500000, 450000, "=F9-G9"],
        [267, "Diğer Maddi Olmayan Hak", "01.02.2024", 0.10, "Hızlı", 900000, 180000, "=F10-G10"],
        [267, "Diğer Duran Varlık", "01.12.2025", 0.20, "Normal", 650000, 65000, "=F11-G11"],
    ]
    _write_table(sheet, 1, headers, examples)
    widths = [14, 25, 18, 16, 18, 18, 20, 18]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, len(examples) + 2):
        for col in range(1, 9):
            cell = sheet.cell(row, col)
            cell.fill = PatternFill("solid", fgColor="E7E6E6")
            if col in (1, 2, 5):
                cell.number_format = "General"
            elif col == 3:
                cell.number_format = "@"
            elif col == 4:
                cell.number_format = "0.00%"
            elif col in (6, 7, 8):
                cell.number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    workbook.save(path)


def create_result_workbook(results: list[dict[str, Any]], output_path: str | Path, islem_yili: int, donem: int, yd_orani: float) -> dict[str, int]:
    workbook = Workbook()
    main = workbook.active
    main.title = "YD ve Amortisman"
    _write_yd_amortisman_sheet(main, results, islem_yili, donem, yd_orani)
    _write_accounting_vouchers_sheet(workbook.create_sheet("Muhasebe Fişleri"), results)
    workbook.save(output_path)
    return {
        "sabit_kiymet_sayisi": len(results),
        "yd_fis_sayisi": len(_group_revaluation_vouchers(results)),
        "amortisman_fis_sayisi": len(_group_depreciation_vouchers(results)),
    }


def _write_title(sheet, title: str, columns: int) -> None:
    sheet.cell(1, 1, title)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = sheet.cell(1, 1)
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")


def _write_table(sheet, start_row: int, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(col)].width = max(14, min(28, len(header) + 4))
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.border = border
            if isinstance(value, datetime):
                cell.number_format = "dd.mm.yyyy"
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"


def _write_yd_amortisman_sheet(sheet, results: list[dict[str, Any]], islem_yili: int, donem: int, yd_orani: float) -> None:
    sheet.merge_cells("A1:Q1")
    sheet["A1"] = "YENİDEN DEĞERLEME VE AMORTİSMAN TABLOSU"
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A2"] = f"İşlem Yılı: {islem_yili}"
    sheet["A3"] = f"Dönem: {_period_label(donem)}"
    sheet["A4"] = f"YD Oranı: %{yd_orani:.4f}"
    for row in (2, 3, 4):
        sheet.cell(row, 1).font = Font(size=11, bold=True)

    headers = [
        "Sabit Kıymet",
        "Açıklama",
        "Aktif Giriş Tarihi",
        "Amort. Oranı",
        "Amort. Yöntemi",
        "Defter Son Değeri",
        "Defter Birikmiş Amort.",
        "Defter Net Değeri",
        "",
        "YD Oranı",
        "YD Sabit Kıymet",
        "YD Birikmiş Amort.",
        "YD Net Değer",
        "",
        "YD Yıllık Amortisman",
        "YD Dönem Amortismanı",
        "Durum",
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(6, col, header)
        if header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    widths = [12, 20, 15, 12, 15, 18, 18, 18, 2, 12, 18, 18, 18, 2, 18, 18, 18]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    for idx, item in enumerate(results, start=7):
        asset = item["asset"]
        values = [
            asset.kiymet_no,
            asset.kiymet_ad,
            asset.tarih.strftime("%d.%m.%Y"),
            asset.amortisman_orani,
            "Hızlı" if asset.yontem == "Azalan" else "Normal",
            asset.maliyet,
            asset.birikmis_amortisman,
            asset.net_deger,
            None,
            item["yd_orani"] / 100,
            item["revalued_cost"],
            item["revalued_accumulated"],
            item["revalued_net"],
            None,
            item["annual_depreciation"],
            item["period_depreciation"],
            item["status"],
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(idx, col, value)
            if col in (4, 10):
                cell.number_format = "0.00%"
            elif col in (6, 7, 8, 11, 12, 13, 15, 16):
                cell.number_format = "#,##0.00"
            elif col == 17 and value:
                cell.font = Font(bold=True, color="9C0006")
                cell.fill = PatternFill("solid", fgColor="FFC7CE")

    total_row = 7 + len(results)
    sheet.cell(total_row, 2, "TOPLAM").font = Font(bold=True)
    for col in (6, 7, 8, 11, 12, 13, 15, 16):
        letter = get_column_letter(col)
        cell = sheet.cell(total_row, col, f"=SUM({letter}7:{letter}{total_row - 1})")
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A7"


def _write_accounting_vouchers_sheet(sheet, results: list[dict[str, Any]]) -> None:
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "MUHASEBE FİŞLERİ"
    sheet["A1"].font = Font(size=14, bold=True)
    for col, width in enumerate([15, 30, 18, 18], start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    row = 3
    row = _write_section_header(sheet, row, "YENİDEN DEĞERLEME FİŞLERİ")
    for account_code, totals in _group_revaluation_vouchers(results).items():
        row = _write_revaluation_voucher(sheet, row, account_code, totals)
        row += 1

    row += 1
    row = _write_section_header(sheet, row, "AMORTİSMAN FİŞLERİ")
    for account_code, amount in _group_depreciation_vouchers(results).items():
        row = _write_depreciation_voucher(sheet, row, account_code, amount)
        row += 1


def _write_section_header(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = sheet.cell(row, 1, title)
    cell.font = Font(size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    return row + 1


def _write_voucher_headers(sheet, row: int) -> None:
    headers = ["Hesap Kodu", "Hesap Adı", "Borç", "Alacak"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")


def _write_revaluation_voucher(sheet, row: int, account_code: str, totals: dict[str, float]) -> int:
    sheet.cell(row, 1, f"Yeniden Değerleme - {account_code}").font = Font(bold=True)
    row += 1
    _write_voucher_headers(sheet, row)
    row += 1
    sheet.cell(row, 1, _account_value(account_code))
    sheet.cell(row, 3, totals["asset_increase"]).number_format = "#,##0.00"
    row += 1
    sheet.cell(row, 1, 257)
    sheet.cell(row, 4, totals["accumulated_increase"]).number_format = "#,##0.00"
    row += 1
    sheet.cell(row, 1, 522)
    sheet.cell(row, 4, totals["fund_increase"]).number_format = "#,##0.00"
    return row + 1


def _write_depreciation_voucher(sheet, row: int, account_code: str, amount: float) -> int:
    sheet.cell(row, 1, f"Dönem Amortismanı - {account_code}").font = Font(bold=True)
    row += 1
    _write_voucher_headers(sheet, row)
    row += 1
    sheet.cell(row, 1, 770)
    sheet.cell(row, 3, amount).number_format = "#,##0.00"
    row += 1
    sheet.cell(row, 1, 257)
    sheet.cell(row, 4, amount).number_format = "#,##0.00"
    return row + 1


def _group_revaluation_vouchers(results: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
    grouped: dict[str, dict[str, float]] = {}
    for item in results:
        if item["fund_increase"] <= 0:
            continue
        account_code = str(item["asset"].aktif_hesap)
        grouped.setdefault(account_code, {"asset_increase": 0.0, "accumulated_increase": 0.0, "fund_increase": 0.0})
        grouped[account_code]["asset_increase"] += item["revaluation_increase"]
        grouped[account_code]["accumulated_increase"] += item["accumulated_increase"]
        grouped[account_code]["fund_increase"] += item["fund_increase"]
    return grouped


def _group_depreciation_vouchers(results: list[dict[str, Any]]) -> dict[str, float]:
    grouped: dict[str, float] = {}
    for item in results:
        if item["period_depreciation"] <= 0:
            continue
        account_code = str(item["asset"].aktif_hesap)
        grouped[account_code] = grouped.get(account_code, 0.0) + item["period_depreciation"]
    return grouped


def _period_label(donem: int) -> str:
    return {1: "1. Dönem", 2: "2. Dönem", 3: "3. Dönem", 4: "Yıllık"}.get(donem, "Yıllık")


def _account_value(account_code: str) -> int | str:
    try:
        return int(account_code)
    except ValueError:
        return account_code


def _write_summary(sheet, results: list[dict[str, Any]], islem_yili: int, donem: int, yd_orani: float) -> None:
    _write_title(sheet, "YENİDEN DEĞERLEME VE AMORTİSMAN ÖZETİ", 4)
    rows = [
        ["İşlem Yılı", islem_yili],
        ["Dönem", {1: "1. Dönem", 2: "2. Dönem", 3: "3. Dönem", 4: "Yıllık"}.get(donem, "Yıllık")],
        ["Yeniden Değerleme Oranı", yd_orani / 100],
        ["Sabit Kıymet Sayısı", len(results)],
        ["Toplam Maliyet", sum(item["asset"].maliyet for item in results)],
        ["Toplam YD Artışı", sum(item["revaluation_increase"] for item in results)],
        ["Dönem Amortismanı", sum(item["period_depreciation"] for item in results)],
    ]
    for row_idx, row in enumerate(rows, start=3):
        sheet.cell(row_idx, 1, row[0]).font = Font(bold=True)
        sheet.cell(row_idx, 2, row[1])
        if row_idx in (5,):
            sheet.cell(row_idx, 2).number_format = "0.00%"
        if row_idx >= 7:
            sheet.cell(row_idx, 2).number_format = "#,##0.00"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 20


def _write_assets_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = ["sabit kıymet", "sabit kıymet açıklama", "aktife giriş tarihi", "amortisman oranı", "amortisman yöntemi", "defter son değeri", "defter birikmiş amort", "defter net değeri"]
    rows = [[item["asset"].kiymet_no, item["asset"].kiymet_ad, item["asset"].tarih, item["asset"].amortisman_orani, item["asset"].yontem, item["asset"].maliyet, item["asset"].birikmis_amortisman, item["asset"].net_deger] for item in results]
    _write_table(sheet, 1, headers, rows)
    for row in range(2, len(rows) + 2):
        sheet.cell(row, 4).number_format = "0.00%"


def _write_revaluation_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = ["Kıymet No", "Kıymet Adı", "Eski Maliyet", "YD Oranı", "YD Artışı", "Yeni Değer"]
    rows = [[item["asset"].kiymet_no, item["asset"].kiymet_ad, item["asset"].maliyet, item["yd_orani"] / 100, item["revaluation_increase"], item["revalued_cost"]] for item in results]
    _write_table(sheet, 1, headers, rows)
    for row in range(2, len(rows) + 2):
        sheet.cell(row, 4).number_format = "0.00%"


def _write_depreciation_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = ["Kıymet No", "Kıymet Adı", "Aktif Ay", "Yıllık Amortisman", "Dönem Amortismanı", "YD Sonrası Dönem Amortismanı"]
    rows = [[item["asset"].kiymet_no, item["asset"].kiymet_ad, item["active_months"], item["annual_depreciation"], item["period_depreciation"], item["revalued_period_depreciation"]] for item in results]
    _write_table(sheet, 1, headers, rows)


def _write_journal_sheet(sheet, results: list[dict[str, Any]]) -> None:
    headers = ["Fiş Tipi", "Kıymet No", "Borç Hesap", "Alacak Hesap", "Tutar", "Açıklama"]
    rows: list[list[Any]] = []
    for item in results:
        asset = item["asset"]
        if item["revaluation_increase"] > 0:
            rows.append(["YD", asset.kiymet_no, asset.aktif_hesap, "522", item["revaluation_increase"], f"{asset.kiymet_ad} yeniden değerleme artışı"])
        if item["period_depreciation"] > 0:
            rows.append(["AMORT", asset.kiymet_no, asset.gider_hesap, "257", item["period_depreciation"], f"{asset.kiymet_ad} dönem amortismanı"])
    _write_table(sheet, 1, headers, rows)
